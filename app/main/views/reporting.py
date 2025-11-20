import jdatetime
from datetime import date
from openpyxl import Workbook
from django.db.models import Sum
from django.utils import timezone
from django.http import JsonResponse
from openpyxl.utils import get_column_letter
from django.utils.encoding import escape_uri_path
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from openpyxl.styles import (
    Font,
    Alignment,
)
from django.shortcuts import (
    render,
    HttpResponse,
)
from ..models import (
    Tag,
    Card,
    Category,
    Transaction,
)
from ..utils import (
    get_jalali_date,
    format_currency,
    format_card_number_last4,
    TRANSACTION_AND_KIND_CHOICES,
)


@login_required
def reporting(request):
    today = date.today()
    jy, jm, jd = get_jalali_date(today).split('/')
    context = {
        'categories': Category.objects.all().order_by('name'),
        'cards': Card.objects.filter(active=True).order_by('name'),
        'tags': Tag.objects.all().order_by('name'),
        'current_day': jd,
        'current_month': jm,
        'current_year': jy,
    }
    return render(request, 'main/reporting.html', context)


@login_required
@require_http_methods(["GET"])
def filter_transactions_ajax(request):
    transaction_type = request.GET.get('type')
    category_id = request.GET.get('category')
    source_card_id = request.GET.get('source')
    destination_card_id = request.GET.get('destination')
    tag_id = request.GET.get('tag')
    report_type = request.GET.get('report_type')
    day = request.GET.get('day')
    month = request.GET.get('month')
    year = request.GET.get('year')

    queryset = Transaction.objects.select_related(
        'category',
        'source',
        'destination',
    ).prefetch_related('tags').all()

    if transaction_type and transaction_type != 'all':
        queryset = queryset.filter(kind=transaction_type)
    if category_id and category_id != 'all':
        queryset = queryset.filter(category_id=category_id)
    if tag_id and tag_id != 'all':
        queryset = queryset.filter(tags__id=tag_id)
    if source_card_id and source_card_id != 'all':
        queryset = queryset.filter(source_id=source_card_id)
    if destination_card_id and destination_card_id != 'all':
        queryset = queryset.filter(destination_id=destination_card_id)

    start_date_gregorian, end_date_gregorian = None, None
    try:
        int_year = int(year) if year else None
        int_month = int(month) if month else None
        int_day = int(day) if day else None

        if report_type == 'daily' and int_year and int_month and int_day:
            jalali_date = jdatetime.date(
                int_year,
                int_month,
                int_day,
            )
            start_date_gregorian = jalali_date.togregorian()
            end_date_gregorian = start_date_gregorian

        elif report_type == 'monthly' and int_year and int_month:
            start_jalali = jdatetime.date(
                int_year,
                int_month,
                1,
            )
            start_date_gregorian = start_jalali.togregorian()

            try:
                end_jalali = jdatetime.date(
                    int_year,
                    int_month,
                    31,
                )
            except ValueError:
                try:
                    end_jalali = jdatetime.date(
                        int_year,
                        int_month,
                        30,
                    )
                except ValueError:
                    end_jalali = jdatetime.date(
                        int_year,
                        int_month,
                        29,
                    )
            end_date_gregorian = end_jalali.togregorian()

        elif report_type == 'annual' and int_year:
            start_jalali = jdatetime.date(
                int_year,
                1,
                1,
            )
            start_date_gregorian = start_jalali.togregorian()

            try:
                end_date_gregorian = jdatetime.date(
                    int_year,
                    12,
                    30,
                ).togregorian()
            except ValueError:
                end_date_gregorian = jdatetime.date(
                    int_year,
                    12,
                    29,
                ).togregorian()

        if start_date_gregorian and end_date_gregorian:
            queryset = queryset.filter(
                date__range=(
                    start_date_gregorian,
                    end_date_gregorian,
                ),
            )
    except (ValueError, TypeError):
        return JsonResponse(
            {
                'transactions': [],
                'error': 'تاریخ وارد شده نامعتبر است.',
            },
            status=400,
        )

    total_amount = None
    if transaction_type and transaction_type != 'all':
        total_sum = queryset.aggregate(total=Sum('amount'))
        if total_sum and total_sum['total'] is not None:
            total_amount = total_sum['total']

    transactions_list = []
    for tx in queryset:
        jalali_dt = get_jalali_date(tx.date)
        amount_formatted = format_currency(tx.amount)

        source_display = 'نامشخص'
        if tx.source:
            if not tx.source.active:
                source_display = (f"{tx.source.get_name_display()} ({tx.source.owner}) - "
                                  f"{format_card_number_last4(tx.source.number)} - غیرفعال")
            else:
                source_display = (f"{tx.source.get_name_display()} ({tx.source.owner}) - "
                                  f"{format_card_number_last4(tx.source.number)}")

        destination_display = 'نامشخص'
        if tx.destination:
            if not tx.destination.active:
                destination_display = (f"{tx.destination.get_name_display()} ({tx.destination.owner}) - "
                                       f"{format_card_number_last4(tx.destination.number)} - غیرفعال")
            else:
                destination_display = (f"{tx.destination.get_name_display()} ({tx.destination.owner}) - "
                                       f"{format_card_number_last4(tx.destination.number)}")

        tag_names = ', '.join(
            [
                tag.name
                for tag in tx.tags.all()
            ]
        )

        transactions_list.append({
            'id': tx.id,
            'date': jalali_dt,
            'kind': tx.kind,
            'kind_display': tx.get_kind_display(),
            'category_name': tx.category.name if tx.category else '---',
            'amount': tx.amount,
            'amount_formatted': amount_formatted,
            'source_card_name': source_display,
            'destination_card_name': destination_display,
            'description': tx.description,
            'tag_names': tag_names,
        })

    return JsonResponse(
        {
            'transactions': transactions_list,
            'total_amount': total_amount,
            'total_amount_formatted': format_currency(total_amount)
            if total_amount is not None else None
        },
    )


@login_required
@require_http_methods(["GET"])
def export_transactions_excel(request):
    transaction_type = request.GET.get('type')
    category_id = request.GET.get('category')
    source_card_id = request.GET.get('source')
    destination_card_id = request.GET.get('destination')
    report_type = request.GET.get('report_type')
    tag_id = request.GET.get('tag')
    day = request.GET.get('day')
    month = request.GET.get('month')
    year = request.GET.get('year')

    queryset = Transaction.objects.select_related(
        'category',
        'source',
        'destination',
    ).prefetch_related('tags').all()

    if transaction_type and transaction_type != 'all':
        queryset = queryset.filter(kind=transaction_type)
    if category_id and category_id != 'all':
        queryset = queryset.filter(category_id=category_id)
    if tag_id and tag_id != 'all':
        queryset = queryset.filter(tags__id=tag_id)
    if source_card_id and source_card_id != 'all':
        queryset = queryset.filter(source_id=source_card_id)
    if destination_card_id and destination_card_id != 'all':
        queryset = queryset.filter(destination_id=destination_card_id)

    start_date_gregorian, end_date_gregorian = None, None
    report_type_display = 'کل-تاریخ‌ها'
    date_part = ''

    try:
        jdate = jdatetime_date if 'jdatetime_date' in globals() else jdatetime.date

        int_year = int(year) if year else None
        int_month = int(month) if month else None
        int_day = int(day) if day else None

        if report_type == 'daily' and int_year and int_month and int_day:
            jalali_date = jdate(
                int_year,
                int_month,
                int_day,
            )
            start_date_gregorian = jalali_date.togregorian()
            end_date_gregorian = start_date_gregorian
            report_type_display = 'روزانه'
            date_part = f"{int_day}-{int_month}-{int_year}"

        elif report_type == 'monthly' and int_year and int_month:
            start_jalali = jdate(
                int_year,
                int_month,
                1,
            )
            start_date_gregorian = start_jalali.togregorian()
            try:
                end_jalali = jdate(
                    int_year,
                    int_month,
                    31,
                )
            except ValueError:
                try:
                    end_jalali = jdate(
                        int_year,
                        int_month,
                        30,
                    )
                except ValueError:
                    end_jalali = jdate(
                        int_year,
                        int_month,
                        29,
                    )
            end_date_gregorian = end_jalali.togregorian()
            report_type_display = 'ماهانه'
            date_part = f"{int_month}-{int_year}"

        elif report_type == 'annual' and int_year:
            start_date_gregorian = jdate(
                int_year,
                1,
                1,
            ).togregorian()
            try:
                end_date_gregorian = jdate(
                    int_year,
                    12,
                    30,
                ).togregorian()
            except ValueError:
                end_date_gregorian = jdate(
                    int_year,
                    12,
                    29,
                ).togregorian()
            report_type_display = 'سالانه'
            date_part = f"{int_year}"

        if start_date_gregorian and end_date_gregorian:
            end_date_with_time = end_date_gregorian + timezone.timedelta(days=1)
            queryset = queryset.filter(
                date__gte=start_date_gregorian,
                date__lt=end_date_with_time,
            )
    except (ValueError, TypeError):
        queryset = Transaction.objects.none()

    total_amount = None
    if transaction_type and transaction_type != 'all':
        total_sum = queryset.aggregate(total=Sum('amount'))
        if total_sum and total_sum['total'] is not None:
            total_amount = total_sum['total']

    transactions = queryset.order_by('-date')

    kind_choices_map = dict(TRANSACTION_AND_KIND_CHOICES)
    file_name_parts = ['گزارش']

    if transaction_type and transaction_type != 'all':
        kind_name = kind_choices_map.get(transaction_type, 'نامشخص')
        file_name_parts.append(kind_name)

    final_report_type_part = report_type_display
    if date_part:
        final_report_type_part += f"-{date_part}"

    file_name_parts.append(final_report_type_part)
    filename = "-".join(file_name_parts) + ".xlsx"
    filename = filename.replace(" ", "_")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'

    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش تراکنش‌ها"
    ws.sheet_view.rightToLeft = True

    columns = [
        'ردیف',
        'تاریخ',
        'نوع تراکنش',
        'مبلغ (تومان)',
        'دسته‌بندی',
        'تگ‌ها',
        'حساب مبدأ',
        'حساب مقصد',
        'توضیحات',
    ]
    ws.append(columns)

    current_row = ws.max_row
    for index, tx in enumerate(transactions, start=1):
        current_row += 1
        source_display = 'نامشخص'
        if tx.source:
            if not tx.source.active:
                source_display = (f"{tx.source.get_name_display()} ({tx.source.owner}) - "
                                  f"{format_card_number_last4(tx.source.number)} - غیرفعال")
            else:
                source_display = (f"{tx.source.get_name_display()} ({tx.source.owner}) - "
                                  f"{format_card_number_last4(tx.source.number)}")

        destination_display = 'نامشخص'
        if tx.destination:
            if not tx.destination.active:
                destination_display = (f"{tx.destination.get_name_display()} ({tx.destination.owner}) - "
                                       f"{format_card_number_last4(tx.destination.number)} - غیرفعال")
            else:
                destination_display = (f"{tx.destination.get_name_display()} ({tx.destination.owner}) - "
                                       f"{format_card_number_last4(tx.destination.number)}")

        jalali_dt = get_jalali_date(tx.date)

        tag_names = ', '.join(
            [
                tag.name
                for tag in tx.tags.all()
            ]
        )

        row = [
            index,
            jalali_dt,
            tx.get_kind_display(),
            float(tx.amount),
            tx.category.name if tx.category else '---',
            tag_names,
            source_display,
            destination_display,
            tx.description,
        ]
        ws.append(row)

    if total_amount is not None:
        current_row += 2

        start_cell_merge = f'A{current_row}'
        amount_cell = f'D{current_row}'
        end_cell_merge = f'C{current_row}'

        ws[start_cell_merge] = 'جمع مبالغ:'
        ws[amount_cell] = float(total_amount)

        ws.merge_cells(f'{start_cell_merge}:{end_cell_merge}')

        bold_font = Font(bold=True)

        ws[start_cell_merge].font = bold_font
        ws[amount_cell].font = bold_font

        ws[start_cell_merge].alignment = Alignment(horizontal='left')

        ws[amount_cell].number_format = '#,##0'

    for col_idx, column in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = len(column) + 5

    wb.save(response)
    return response